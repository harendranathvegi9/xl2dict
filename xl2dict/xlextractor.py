# Copyright (c) 2016 Ashwin Kondapalli
#
# See the file LICENSE.txt for copying permission.
__author__ = 'ashwin'
from xlrd import open_workbook
import itertools
import os
from openpyxl import load_workbook


class XlToDict():
    def convert_sheet_to_dict(self, file_path=None, sheet=None, filter_variables_dict=None):
        """
        This method will convert excel sheets to dict. The input is path to the excel file or a sheet object.
        if file_path is None, sheet object must be provded. This method will convert only the first sheet.
        If you need to convert multiple sheets, please use the method fetch_data_by_column_by_sheet_name_multiple() and
        fetch_data_by_column_by_index_multiple().If you need to filter data by a specific keyword, specify the dict in
        filter_variables_dict like {column name : keyword} . Any rows that matches the keyword in the specified column
        will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet: The sheet object from the spreadsheet
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name : keyword}
        :return result_dict_list: List of dictionaries .
        """
        # read headers
        if file_path is not None:
            workbook = open_workbook(filename=file_path)
            sheet = workbook.sheet_by_index(0)
        keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]
        # print keys
        found_row_dict_list = []
        for column_index, key in enumerate(keys):
            if filter_variables_dict is not None:
                for column_name, column_value in filter_variables_dict.iteritems():
                    if key == column_name:
                        for row_index in xrange(sheet.nrows):
                            if not (column_name == None and column_value == None):
                                if (sheet.cell(row_index, column_index).value) == column_value:
                                    found_row_dict = {
                                        keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                        for col_index_internal in xrange(sheet.ncols)}
                                    found_row_dict_list.append(found_row_dict)
                            else:
                                found_row_dict = {
                                    keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                    for col_index_internal in xrange(sheet.ncols)}
                                found_row_dict_list.append(found_row_dict)
            elif filter_variables_dict == {} or filter_variables_dict is None:
                filter_variables_dict = {}
                for row_index in xrange(sheet.nrows):
                    found_row_dict = {keys[col_index_internal]: sheet.cell(row_index, col_index_internal).value
                                      for col_index_internal in xrange(sheet.ncols)}
                    found_row_dict_list.append(found_row_dict)
                del found_row_dict_list[0]
        result_dict_list = []
        if len(found_row_dict_list) > 1 and len(filter_variables_dict) > 1:
            for a, b in itertools.combinations(found_row_dict_list, len(filter_variables_dict)):
                if a == b:
                    result_dict_list.append(a)
        else:
            result_dict_list = found_row_dict_list
        return result_dict_list

    def fetch_data_by_column_by_sheet_name(self, file_path, filter_variables_dict=None, sheet_name=None):
        """
        This method will convert the specified sheet in the excel file to dict. The input is path to the excel file .
        If sheet_name is not provided, this method will convert only the first sheet.
        If you need to convert multiple sheets, please use the method fetch_data_by_column_by_sheet_name_multiple() or
        fetch_data_by_column_by_sheet_index_multiple(). If you need to filter data by a specific keyword,
        specify the dict in filter_variables_dict like {column name : keyword} . Any rows that matches the keyword in
        the specified column will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_name: Optional. The name of the sheet in the spreadsheet to convert. If not specified, will convert
         the first sheet
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return result_dict_list: List of dictionaries .
        """
        workbook = open_workbook(filename=file_path)
        if sheet_name is not None:
            sheet = workbook.sheet_by_name(sheet_name)
        else:
            sheet = workbook.sheet_by_index(0)
        return self.convert_sheet_to_dict(sheet, filter_variables_dict)

    def fetch_data_by_column_by_sheet_index(self, file_name, filter_variables_dict=None, sheet_index=0):
        """
        This method will convert the specified sheet in the excel file to dict. The input is path to the excel file .
        If sheet_index is not provided, this method will convert only the first sheet.
        If you need to convert multiple sheets, please use the method fetch_data_by_column_by_sheet_name_multiple() or
        fetch_data_by_column_by_sheet_index_multiple(). If you need to filter data by a specific keyword,
        specify the dict in filter_variables_dict like {column name : keyword} . Any rows that matches the keyword in
        the specified column will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_index: Optional. The index of the sheet in the spreadsheet to convert. If not specified, will
        convert the first sheet
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return result_dict_list: List of dictionaries .
        """
        workbook = open_workbook(filename=file_name)
        sheet = workbook.sheet_by_index(sheet_index)
        return self.convert_sheet_to_dict(sheet, filter_variables_dict)

    def fetch_data_by_column_by_sheet_name_multiple(self, file_name, filter_variables_dict=None, sheet_names=None):
        """
        This method will convert multiple sheets in the excel file to dict. The input is path to the excel file .
        If sheet_names is not provided, this method will convert ALL the sheets.If you need to filter data by a specific
        keyword / keywords, specify the dict in filter_variables_dict like {column name : keyword} .
        Any rows that matches the keyword  in the specified column will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_names: Optional. The list of sheet names in the spreadsheet to convert. If not specified, will
        convert all the sheets
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return result_dict_list: List of dictionaries .
        """
        workbook = open_workbook(filename=file_name)
        resultdictlist = []
        if sheet_names is None:
            sheet_names = workbook.sheet_names()
            print sheet_names
        for sheet_name in sheet_names:
            sheet = workbook.sheet_by_name(sheet_name)
            resultdictlist.extend(self.convert_sheet_to_dict(sheet, filter_variables_dict))
        return resultdictlist

    def fetch_data_by_column_by_sheet_index_multiple(self, file_name, filter_variables_dict=None,
                                                     sheet_indices=None):
        """
        This method will convert multiple sheets in the excel file to dict. The input is path to the excel file .
        If sheet_indices is not provided, this method will convert ALL the sheets.If you need to filter data by a
        specific keyword / keywords, specify the dict in filter_variables_dict like {column name : keyword} .
        Any rows that matches the keyword  in the specified column will be returned. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_indices: Optional. The list of sheet indices in the spreadsheet to convert. If not specified,
        will convert all the sheets
        :param filter_variables_dict: Optional. The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return result_dict_list: List of dictionaries .
        """
        workbook = open_workbook(filename=file_name)
        resultdictlist = []
        if sheet_indices is None:
            sheet_indices = range(0, workbook.nsheets)
        for sheet_index in sheet_indices:
            sheet = workbook.sheet_by_index(sheet_index)
            resultdictlist.extend(self.convert_sheet_to_dict(sheet, filter_variables_dict))
        return resultdictlist

    def fetch_matching_data_row_indices(self, file_name, filter_variables_dict, sheet_name_index=0):
        """
        This method will fetch all the rows matching the specified filter. The input is path to the excel file .
        If sheet_name_index is not provided, this method will search the first sheet sheet. If you need to filter data
        by a specific keyword / keywords, specify the dict in filter_variables_dict like {column name : keyword} .
        All the row indices that matches the keyword  in the specified column will be returned. Multiple keywords can be
         specified.

        :param file_path: The path to the spreadsheet
        :param sheet_name_index: Optional. The Index or Name of the sheet in the spreadsheet to search. If not
        specified, will search in the first sheet
        :param filter_variables_dict: The keyword dict to filter out the data. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return match_list: List of row indices that matches the filter .
        """
        workbook = open_workbook(filename=file_name)
        if type(sheet_name_index) is int:
            sheet = workbook.sheet_by_index(sheet_name_index)
        else:
            sheet = workbook.sheet_by_name(sheet_name_index)
        # read headers
        keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]
        unfound_match_list = []
        match_list = []
        for column_index, key in enumerate(keys):
            for column_name, column_value in filter_variables_dict.iteritems():
                if key == column_name:
                    for row_index in xrange(sheet.nrows):
                        if not (column_name == None and column_value == None):
                            if str((sheet.cell(row_index, column_index).value)).lower() == column_value.lower():
                                unfound_match_list.append(row_index)
        if len(unfound_match_list) > 1 and len(filter_variables_dict) > 1:
            for a, b in itertools.combinations(unfound_match_list, len(filter_variables_dict)):
                if a == b:
                    match_list.append(a)
        else:
            match_list = unfound_match_list
        return match_list

    def write_data_to_column(self, file_name, column_name, data, filter_variables_dict=None,
                             sheet_name=None):
        """
        This method will write data in to the specified column of all the rows matching the specified filter. The input
        is path to the excel file .If sheet_name is not provided, this method will write data in to the specified column
         in the first sheet sheet. If you need to write data  in to rows by a specific keyword / keywords, specify the
         dict in filter_variables_dict like {column name : keyword} .The specified data will be written in the specified
          column in all rows that matches the keyword. Multiple keywords can be specified.

        :param file_path: The path to the spreadsheet
        :param sheet_name: Optional. The name of the sheet in the spreadsheet to write data into. If not specified,
        data will be written in the first sheet
        :param column_name:The name of the column in the spreadsheet to write data into.
        :param data:The data to write.
        :param filter_variables_dict: Optional. The keyword dict to filter out the rows to write into. specified as
        {column name-1 : keyword1, column name-2 : keyword2}
        :return will return 1 if the write is successful .
        """
        workbook = open_workbook(filename=file_name)
        if sheet_name is None:
            sheet_name = workbook.sheet_names()[0]
        sheet = workbook.sheet_by_name(sheet_name)
        # read headers
        keys = [sheet.cell(0, col_index).value for col_index in xrange(sheet.ncols)]
        column_index = keys.index(column_name)
        if filter_variables_dict is None:
            row_indices_list = range(1, sheet.nrows)
        else:
            row_indices_list = self.fetch_matching_data_row_indices(file_name, filter_variables_dict,
                                                                    sheet_name_index=sheet_name)
        workbook_write = load_workbook(filename=file_name)
        sheet_write = workbook_write.get_sheet_by_name(sheet_name)
        for row_index in row_indices_list:
            sheet_write.cell(row=row_index + 1, column=column_index + 1).value = data
        os.remove(file_name)
        workbook_write.save(file_name)
        return 1
